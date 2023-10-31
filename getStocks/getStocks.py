# /opt/homebrew/bin/python3
'''
getStocks.py: This is my first attempt at a useful Python script.  This
script uses yFinance to lookup some of the stocks that I own and will
get the closing price.  It will then save the data to an Excel spreadsheet.
The script will not fetch stock prices while the Nasdaq and NYSE are open.
It will also not fetch data on the weekends.

    Requires:
       yfinance
       tkinter
       csv
       openpyxl
       datetime
       zoneinfo

    Methods:
        read_stock_file()
        getChange()
        getPercentChange()
        getLastRow()
        createSheets()
        addSheet()
        createColumnNames()
        formatHeader()
        isMarketOpen()
        getMarketOpenHours()
        hasScriptBeenRunToday()
        openExcelFile()
        checkStocks()
        main()

    To Do: 
    1) Add UI to add/remove stocks?
    2) Add graphing of data (once alot more data has been collected)
'''

import yfinance as yf
from tkinter import messagebox
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

# Global Variables
myStocks = []
myFile = 'myStocks.xlsx'

# readStockFile()- Get list of stocks from external .csv file stocktickers.csv
# Requires:
#   file- 'stockticker.csv'.  A .csv file with company ticker, name
#
# Returns:
#
def readStockFile():
    try:
        with open('stocktickers.csv', 'r') as read_obj:
            csv_reader = csv.reader(read_obj)
            stockTickers = list(csv_reader)
        for cticker in stockTickers:
            myStocks.append(cticker[0])
    except FileNotFoundError as e:
        print(e)
    except:
        print('Something went wrong with accessing file stocktickers.csv')

# getChange()- Calculate the difference between the previous closing price and today's closing price
# Requires:
#   curPrice- Today's closing price
#   oldPrice- Previous closing price
#
# Returns:
#   change- Float of the change in price
#
def getChange(curPrice, oldPrice):
    change = float(curPrice - oldPrice)
    return change

# getPercentChange()- Calculate the percentage change between the previous closing price and today's closing price
# Requires:
#   curPrice- Today's closing price
#   oldPrice- Previous closing price
#
# Returns:
#   pchange- Float of teh change in price
#
def getPercentChange(curPrice, oldPrice):
    pchange = float(((curPrice - oldPrice) / oldPrice) * 100)
    return pchange

# getLastRow()- Get index of the first empty row for each sheet in workbook for the Excel file
#
# Requires:
#   sheet- the Workbook object
#
# Returns:
#   index of the last row of the workbook (plus 1, the first empty row)
#
def getLastRow(sheet):
    last_row = sheet.max_row
    last_col = sheet.max_column
    
    for i in range(last_row):
        for j in range(last_col):
            if sheet.cell(last_row, last_col).value is None:
                last_row -= 1
                last_col -= 1 
    return last_row+1

# createSheets()- Using the list of stock tickers in myStocks, create a sheet for each stock
#
# Requires: 
#   wb_obj- The Excel workbook object
#
# Returns:
#
def createSheets(wb_obj):
    # Create the sheets
    for stocks in myStocks:
        if stocks == 'AAPL':
            sheet = wb_obj.active
            sheet.title = stocks
        else:
            sheet = wb_obj.create_sheet(title=stocks)

# addSheet()- Add new sheet for stock symbol that may be missing from workbook
#
# Requires: 
#   wb_obj- The Excel workbook object
#   symb-   The stocks symbol
#
# Returns:
#   sheet- workbook sheet
#           
def addSheet(wb_obj, symb):
    sheet = wb_obj.create_sheet(title=symb)
    return sheet

# createColumnNames()- Create the column names for each sheet in the Excel workbook
#
# Requires: 
#   wb_obj- The Excel workbook object
#
# Returns:
#
def createColumnNames(wb_obj):
    # Create the column names in each sheet and set their width
    sheetNames = wb_obj.sheetnames
    for name in sheetNames:
        sheet = wb_obj[name]
        sheet['A1'] = 'Date'
        sheet.column_dimensions['A'].width = 12
        sheet['B1'] = 'Ticker'
        sheet.column_dimensions['B'].width = 10
        sheet['C1'] = 'Closing Price'
        sheet.column_dimensions['C'].width = 15
        sheet['D1'] = 'Previous Close'
        sheet.column_dimensions['D'].width = 15
        sheet['E1'] = 'Change'
        sheet.column_dimensions['E'].width = 12
        sheet['F1'] = '% Change'
        sheet.column_dimensions['F'].width = 12

# formatHeader()- Format the first row on (column header) for each sheet in the Excel workbook
#
# Requires: 
#   wb_obj- The Excel workbook object
#
# Returns:
#
def formatHeader(wb_obj):
    # Format the top row on each sheet
    lightGrey = 'CCCCCCCC'
    sheetNames = wb_obj.sheetnames
    for name in sheetNames:
        sheet = wb_obj[name]
        for columns in sheet.iter_cols(min_col=1, max_col=6):
            for cell in columns:
                cell.fill = PatternFill(start_color=lightGrey, end_color=lightGrey,
                                        fill_type = "solid")
                cell.font = Font(bold = True)
                cell.alignment = Alignment(horizontal = 'center')

# isMarketOpen()- Check to see if the NYSE/NASDAQ is open or closed. Script will only run if NYSE/NASDAQ is closed
#               so that I am only dealing with the final closing price.  Script also will not execute on weekends
#               when the stock markets are closed to trading.
#
# Requires:
#
# Returns:
#   Boolean or String-  If markets are open return True.  If closed return False.  If it is the weekend return string 'weekend'
def isMarketOpen():
    utcnow = datetime.now(timezone.utc)
    nytz = ZoneInfo('America/New_York')
    nynow = utcnow.astimezone(nytz)
    nywday = nynow.isoweekday()
    if nywday > 0 and nywday < 6:
        utcnow_hour_min = utcnow.strftime('%H:%M')
        if (((utcnow_hour_min > '13:30')) and (utcnow.hour < 20)):
            return True
        else:
            return False
    else:
        return 'weekend'

# getMarketOpenHours()- Get the hours the NYSE and NASDAQ are open and return a list with time_open, time_close
#               and timezone strings for user's local timezone
# 
# Required: 
#
# Returns:
#   List of strings with the market open time, market close time and timezone for the local timezone
def getMarketOpenHours():
    utc_time = datetime.now(timezone.utc)
    utc_start = utc_time.replace(hour=13, minute=30, second=0, microsecond=0)
    utc_end = utc_time.replace(hour=20, minute=0, second=0, microsecond=0)
    localTZ = datetime.now(timezone.utc).astimezone().tzinfo
    local_start = utc_start.astimezone(localTZ)
    local_end = utc_end.astimezone(localTZ)
    return [local_start.strftime('%I:%M %p'), local_end.strftime('%I:%M %p'), localTZ]

# hasScriptBeenRunToday()- Check to see if script was already run.  Looks for last date in excel workbook.
#
# Requires:
#
# Returns: 
#   Boolean
def hasScriptBeenRunToday():
    try:
        workbook = openpyxl.load_workbook(myFile)
        wp = workbook.active
        last_row = wp.max_row
        todayDate = datetime.now().strftime('%Y-%m-%d')
        prevDate = wp.cell(last_row, 1).value
        if todayDate == prevDate:
            # Script has already been run today
            return True
        else:
            return False
    except FileNotFoundError as e:
        return False
    except:
        print(e)

# openExcelFile()- Create/Open the Excel file, format it if new, then return the workbook object
#
# Requires:
#
# Returns:
#   Excel wookbook object
def openExcelFile():
    # Create/open xlsx file for storing data
    try:
        # Open workbook
        wb_obj = openpyxl.load_workbook(myFile)

    except FileNotFoundError as e:
        wb_obj = Workbook()
        createSheets(wb_obj)
        createColumnNames(wb_obj)
        formatHeader(wb_obj)
 
    except UnboundLocalError as e:
        print(e)
    
    return wb_obj

# checkStocks()- Using list myStocks and yfinance, get stock data on each stock and write it out to file
#
# Requires:
#   wb_obj- Excel workbook object
#   datestr- A date string formatted to %Y-%m-%d, the date format required by yfinance
#
# Returns:
#
def checkStocks(wb_obj, datestr):
    for stocks in myStocks:
        ticker = yf.Ticker(stocks)
        prices = ticker.history(period='2d')
        curr_symbol = stocks

        # For reasons unknown to me, XLU does not have currentPrice as part of Yahoo Finance ticker info.
        # Use navPrice instead
        if curr_symbol == 'XLU':
            current_price = prices.iloc[1,3]
        else:
            current_price = prices.iloc[1,3]
        previous_close_price = prices.iloc[0,3]
        change = getChange(current_price, previous_close_price)
        pctChange = getPercentChange(current_price, previous_close_price)
        # Check if there is a sheet in the workbook for the current stock.  If not, add the sheet and add/format the header
        try:
            sheet = wb_obj[curr_symbol]
        except KeyError:
            sheet = addSheet(wb_obj, curr_symbol)
            createColumnNames(wb_obj)
            formatHeader(wb_obj)
        finally:
            lastRow = getLastRow(sheet)
            cell = sheet['A' + str(lastRow)]
            cell.value = datestr
            cell.number_format = 'yyyy-mm-dd'
            cell.alignment = Alignment(horizontal = 'right')
            cell = sheet['B' + str(lastRow)]
            cell.value = curr_symbol
            cell.alignment = Alignment(horizontal='center')
            cell = sheet['C' + str(lastRow)]
            cell.value = current_price
            cell = sheet['D' + str(lastRow)]
            cell.value = previous_close_price
            cell = sheet['E' + str(lastRow)]
            cell.value = change
            cell.number_format = '0.00'
            cell = sheet['F' + str(lastRow)]
            cell.value = pctChange
            cell.number_format = '0.00'

# main()- Self explanitory
#
# Requires:
#
# Returns:
#
def main():
    # Get current date string formatted my way in user's timezone
    datestr = datetime.now().strftime('%Y-%m-%d')

    # Check if NYSE/NASDAQ are open
    marketOpen = isMarketOpen()
    if marketOpen == True:
        mktHours = getMarketOpenHours()
        mrktOpenMsg = ('NYSE and NASDAQ are currently open.\nPlease wait for these exchanges to close.\nMarket hours are:\nMon - Fri\n{0} - {1} {2}').format(mktHours[0],mktHours[1],mktHours[2])
        print('\a') #Beep sound is played
        messagebox.showwarning('Markets still open', mrktOpenMsg)
        
    else:
        if marketOpen == 'weekend':
            mktHours = getMarketOpenHours()
            mktClosedMsg = ('NYSE and NASDAQ are closed on weekends.\nMarket hours are:\nMon - Fri\n{0} - {1} {2}').format(mktHours[0],mktHours[1],mktHours[2])
            print('\a') #Beep sound is played
            messagebox.showwarning('It\'s the weekend!',  mktClosedMsg)
        else:
            # Check to see if the script was already run today
            alreadyRun = hasScriptBeenRunToday()
            if alreadyRun == False:
                # Get stock tickers
                readStockFile()
                # Open excel file
                myWorkBook = openExcelFile()
                # Get stock quotes
                checkStocks(myWorkBook, datestr)
                # Save the file
                myWorkBook.save(myFile)
            else:
                print('\a') #Beep sound is played
                messagebox.showerror('Script has already run', 'This script was already run today.\nNo action will be taken.')

if __name__ == "__main__":
    main()
