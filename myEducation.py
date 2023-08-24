'''Script for testing out code before incorporating it into getStocksNew.py'''

from datetime import datetime, timezone
from time import gmtime
import time
from tkinter import messagebox
import openpyxl

def last_active_row():
    workbook = openpyxl.load_workbook('myStocks.xlsx')
    wp = workbook['AAPL']
    last_row = wp.max_row
    last_col = wp.max_column
    
    for i in range(last_row):
        for j in range(last_col):
            if wp.cell(last_row, last_col).value is None:
                last_row -= 1
                last_col -= 1 
            else:
                print(wp.cell(last_row,last_col).value) 
    print("The Last active row is: ", (last_row+1)) # +1 for index 0

def getLastStockEntry():
    workbook = openpyxl.load_workbook('myStocks.xlsx')
    wp = workbook.active
    last_row = wp.max_row
    todayDate = datetime.now().strftime('%Y-%m-%d')
    prevDate = wp.cell(last_row, 1).value
    if todayDate == prevDate:
        messagebox.showerror('Dates Match', 'This script was already run today.\nNo action will be taken.')
    else:
        messagebox.showinfo('Good to Go', 'Script is good to go')

def isMarketOpen():
    # Get current date
    utcnow = datetime.now(timezone.utc)
    utcmkt_open = '13:30'
    utcnow_hm = '{0}:{1}'.format(utcnow.hour, utcnow.minute)
    if (((utcnow_hm > utcmkt_open)) and (utcnow.hour < 20)):
        messagebox.showwarning('Warning', 'Markets are still open')
        return False
    else:
        return True
    # return False
    
def getMarketTime():
    utc_time = datetime.now(timezone.utc)
    utc_start = utc_time.replace(hour=13, minute=30, second=0, microsecond=0)
    utc_end = utc_time.replace(hour=20, minute=0, second=0, microsecond=0)
    localTZ = datetime.now(timezone.utc).astimezone().tzinfo
    local_start = utc_start.astimezone(localTZ)
    local_end = utc_end.astimezone(localTZ)
    # print('utc_start:', utc_start, '\nlocal_start:', local_start,'\nutc_end:', utc_end, '\nlocal_end:', local_end)
    return [local_start.strftime('%I:%M %p'), local_end.strftime('%I:%M %p'), localTZ]

def main():
    # last_active_row()
    # marketClosed = isMarketOpen()
    marketClosed = isMarketOpen()
    if marketClosed == True:
        getLastStockEntry()
    else:
        mktHours = getMarketTime()
        mrktOpenMsg = ('NYSE and NASDAQ are still open.\nPlease wait for markets to close.\nMarket hours are:\nMon - Fri\n{0} - {1} {2}').format(mktHours[0],mktHours[1],mktHours[2])
        messagebox.showwarning('Markets still open', mrktOpenMsg)

if __name__ == '__main__':
    main()