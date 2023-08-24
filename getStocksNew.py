# /opt/homebrew/bin/python3
'''
getStocks: This is my first attempt at a useful Python script.  This
script uses yFinance to lookup some of the stocks that I own and will
get the closing price.  It will then save the data to an Excel spreadsheet.

    To Do: 
    1) Read in stocks from external file 
    2) Add UI to add/remove stocks
'''

import yfinance as yf
from tkinter import messagebox
from datetime import datetime, timezone
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# My Stocks
myStocks = ['AAPL', 'ADBE', 'CSCO', 'NTAP', 'MSFT', 'AMZN', 'TSLA', 'VMW', 'DOCU', 'SPLK', 'XLU']

# Get current datetime
now = datetime.now()
utcnow = datetime.now(timezone.utc)

# File to work with
myFile = 'myStocks.xlsx'

# Calculate the difference between today's price and yesterday's price
def getChange(curPrice, oldPrice):
    change = float(curPrice - oldPrice)
    return change

# Calulate the pecent change from today's price and yesterday's price
def getPercentChange(curPrice, oldPrice):
    pchange = float(((curPrice - oldPrice) / oldPrice) * 100)
    return pchange

# Get the index of the first empty row for each sheet in workbook for the Excel file
def getLastRow(sheet):
    last_row = sheet.max_row
    last_col = sheet.max_column
    
    for i in range(last_row):
        for j in range(last_col):
            if sheet.cell(last_row, last_col).value is None:
                last_row -= 1
                last_col -= 1 
    return last_row+1

# Using myStocks, create a sheet for each stock symbol
def createSheets(wb_obj):
    # Create the sheets
    for stocks in myStocks:
        if stocks == 'AAPL':
            sheet = wb_obj.active
            sheet.title = stocks
        else:
            sheet = wb_obj.create_sheet(title=stocks)

# Create the column titles I want and format them
def createColumnNames(wb_obj):
    # Create the column names in each sheet and set their width
    sheetNames = wb_obj.sheetnames
    for name in sheetNames:
        sheet = wb_obj[name]
        sheet['A1'] = 'Date'
        sheet.column_dimensions['A'].width = 10
        sheet['B1'] = 'Ticker'
        sheet.column_dimensions['B'].width = 10
        sheet['C1'] = 'Closing Price'
        sheet.column_dimensions['C'].width = 12
        sheet['D1'] = 'Previous Close'
        sheet.column_dimensions['D'].width = 14
        sheet['E1'] = 'Change'
        sheet.column_dimensions['E'].width = 12
        sheet['F1'] = '% Change'
        sheet.column_dimensions['F'].width = 12

# Format the column headers
def formatHeader(wb_obj):
    # Format the top row on each sheet
    lightGrey = 'CCCCCCCC'
    sheetNames = wb_obj.sheetnames
    for name in sheetNames:
        sheet = wb_obj[name]
        for columns in sheet.iter_cols(min_col=1, max_col=6):
            for cell in columns:
                cell.fill = PatternFill(start_color=lightGrey, end_color=lightGrey,
                                        fill_type = "solid")
                cell.font = Font(bold = True)
                cell.alignment = Alignment(horizontal = 'center')

# Check to see if the NYSE/NASDAQ is open or closed. Script only run if NYSE/NASDAQ is closed
def isMarketOpen():
    # Get current UTC date/time
    utcnow = datetime.now(timezone.utc)
    if (((utcnow.hour > 13) and (utcnow.minute > 30)) and (utcnow.hour < 20)):
        # messagebox.showwarning('Warning', 'Markets are still open')
        return True
    else:
        return False

# Get the hours the NYSE and NASDAQ are open and return 
# a list with time_open, time_close and timezone strings for local timezone
def getMarketOpenHours():
    utc_time = datetime.now(timezone.utc)
    utc_start = utc_time.replace(hour=13, minute=30, second=0, microsecond=0)
    utc_end = utc_time.replace(hour=20, minute=0, second=0, microsecond=0)
    localTZ = datetime.now(timezone.utc).astimezone().tzinfo
    local_start = utc_start.astimezone(localTZ)
    local_end = utc_end.astimezone(localTZ)
    return [local_start.strftime('%I:%M %p'), local_end.strftime('%I:%M %p'), localTZ]

# Check to see if script was already run.  Looks for last date in excel workbook.
def hasScriptBeenRunToday():
    workbook = openpyxl.load_workbook(myFile)
    wp = workbook.active
    last_row = wp.max_row
    todayDate = datetime.now().strftime('%Y-%m-%d')
    prevDate = wp.cell(last_row, 1).value
    if todayDate == prevDate:
        # Script has already been run today
        return True
    else:
       return False

def getStockData():
    # Get current date string formatted my way in user's timezone
    datestr = now.strftime('%Y-%m-%d')

    # Create/open xlsx file for storing data
    try:
        # Open workbook
        wb_obj = openpyxl.load_workbook(myFile)
        checkStocks(wb_obj, datestr)
        
    except FileNotFoundError as e:
        wb_obj = Workbook()
        createSheets(wb_obj)
        createColumnNames(wb_obj)
        formatHeader(wb_obj)
        checkStocks(wb_obj, datestr)
 
    except UnboundLocalError as e:
        print(e)
    
    # Save the file
    wb_obj.save(myFile)

# Using myStocks, get stock data on each stock and write it out to file
def checkStocks(wb_obj, datestr):
    for stocks in myStocks:
        ticker = yf.Ticker(stocks).info
        curr_symbol = ticker['symbol']
        # For reasons unknown to me, XLU does not have currentPrice as part of Yahoo Finance ticker info.
        # Use navPrice instead
        if curr_symbol == 'XLU':
            current_price = ticker['navPrice']
        else:
            current_price = ticker['currentPrice']
        previous_close_price = ticker['regularMarketPreviousClose']
        change = getChange(current_price, previous_close_price)
        pctChange = getPercentChange(current_price, previous_close_price)
        sheet = wb_obj[curr_symbol]
        lastRow = getLastRow(sheet)
        sheet['A' + str(lastRow)] = datestr
        sheet['B' + str(lastRow)] = curr_symbol
        sheet['C' + str(lastRow)] = current_price
        sheet['D' + str(lastRow)] = previous_close_price
        sheet['E' + str(lastRow)] = change
        sheet['F' + str(lastRow)] = pctChange

def main():
    # Check if NYSE/NASDAQ are open
    marketOpen = isMarketOpen()
    if marketOpen == True:
        mktHours = getMarketOpenHours()
        mrktOpenMsg = ('NYSE and NASDAQ are still open.\nPlease wait for markets to close.\nMarket hours are:\nMon - Fri\n{0} - {1} {2}').format(mktHours[0],mktHours[1],mktHours[2])
        messagebox.showwarning('Markets still open', mrktOpenMsg)
    else:
        # messagebox.showinfo('Dialog', 'Market is closed')
        alreadyRun = hasScriptBeenRunToday()
        if alreadyRun == False:
            getStockData()
        else:
            messagebox.showerror('Script has already run', 'This script was already run today.\nNo action will be taken.')

if __name__ == "__main__":
    main()
